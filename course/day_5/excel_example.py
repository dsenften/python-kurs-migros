import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
"""
Mit dieser Funktion lesen wir die Excel Datei ein,
welche im aktuellen Verzeichnis gespeichert sein muss.
"""

type(wb)       # Ausgeben der geladenen Klasse
# <class 'openpyxl.workbook.workbook.Workbook'>

wb.sheetnames  # Liefert eine Liste der Blätter
# ['Tabelle1']

sheet = wb['Tabelle1'] # Auswählen eines konkreten Blattes
# <class 'openpyxl.worksheet.worksheet.Worksheet'>

cell = sheet['B4']     # Auswählen einer einzelnen Zelle
# <Cell 'Tabelle1'.B4>

sheet['B4'].value      # Ausgeben einer einzelnen Zelle

# Viel einfacher kann dies mit einer Funktion erreicht werden
sheet.cell(row=4, column=2).value

# Jetzt geben wir alle Daten einer Spalte (B) aus
for i in range(2, sheet.max_row +1):
    print(sheet.cell(row=i, column=3).value)